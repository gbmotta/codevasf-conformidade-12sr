"""
Análise em lote de vários ZIPs → planilha consolidada.
"""

from __future__ import annotations

import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from conformidade.analyzer import analisar_conformidade
from conformidade.checklist import TipoEntidade, label_tipo, load_checklist
from conformidade.config import Settings
from conformidade.history import save_analysis
from conformidade.inventory_ui import build_inventory
from conformidade.loaders import load_from_zip
from conformidade.ml.cnpj_check import cross_check_cnpj
from conformidade.models import RelatorioConformidade, StatusConformidade


@dataclass
class BatchRow:
    ordem: int
    zip_name: str
    beneficiario: str
    tipo: str
    atendidos: int
    parciais: int
    nao_atendidos: int
    situacao: str
    cnpj: str
    alertas: str
    history_id: str
    erro: str = ""


@dataclass
class BatchResult:
    rows: list[BatchRow] = field(default_factory=list)
    generated_at: str = ""

    def to_xlsx_bytes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Controle"
        headers = [
            "ITEM",
            "ZIP",
            "BENEFICIÁRIO",
            "TIPO",
            "ATENDIDOS",
            "PARCIAIS",
            "NÃO ATENDIDOS",
            "SITUAÇÃO DO PROCESSO",
            "CNPJ",
            "ALERTAS",
            "HISTORY_ID",
            "ERRO",
        ]
        header_fill = PatternFill("solid", fgColor="005CA8")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        fills = {
            "OK": PatternFill("solid", fgColor="D7F0E5"),
            "PENDÊNCIAS": PatternFill("solid", fgColor="EEF7D4"),
            "INCOMPLETO": PatternFill("solid", fgColor="FDE2E2"),
            "ERRO": PatternFill("solid", fgColor="F5C6CB"),
        }
        for r, row in enumerate(self.rows, start=2):
            values = [
                row.ordem,
                row.zip_name,
                row.beneficiario,
                row.tipo,
                row.atendidos,
                row.parciais,
                row.nao_atendidos,
                row.situacao,
                row.cnpj,
                row.alertas,
                row.history_id,
                row.erro,
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(r, c, val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            sit_fill = fills.get(row.situacao.split("—")[0].strip())
            if sit_fill:
                ws.cell(r, 8).fill = sit_fill

        widths = [6, 28, 36, 14, 12, 12, 14, 22, 20, 40, 22, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
        # Fix column letters for >26 — use openpyxl utility
        from openpyxl.utils import get_column_letter

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Resumo")
        ws2["A1"] = "Gerado em"
        ws2["B1"] = self.generated_at or datetime.now().strftime("%d/%m/%Y %H:%M")
        ws2["A2"] = "Total de pacotes"
        ws2["B2"] = len(self.rows)
        ok = sum(1 for r in self.rows if r.situacao.startswith("OK"))
        pend = sum(1 for r in self.rows if r.situacao.startswith("PEND"))
        inc = sum(1 for r in self.rows if r.situacao.startswith("INCOMP"))
        err = sum(1 for r in self.rows if r.situacao.startswith("ERRO"))
        ws2["A3"] = "OK"
        ws2["B3"] = ok
        ws2["A4"] = "Pendências"
        ws2["B4"] = pend
        ws2["A5"] = "Incompletos"
        ws2["B5"] = inc
        ws2["A6"] = "Erros"
        ws2["B6"] = err

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _situacao(rel: RelatorioConformidade) -> str:
    c = rel.contagem
    if c["nao_atendido"] == 0 and c["parcial"] == 0:
        return "OK — documentação aparente completa"
    if c["nao_atendido"] == 0 and c["parcial"] > 0:
        return "PENDÊNCIAS — itens parciais / validade"
    if c["atendido"] == 0:
        return "INCOMPLETO — poucos/nenhum item atendido"
    return "PENDÊNCIAS — faltam documentos"


def analyze_zip_batch(
    settings: Settings,
    zip_paths: list[Path],
    tipo: TipoEntidade,
    *,
    save_history: bool = True,
    work_dir: Path | None = None,
    on_progress=None,
) -> BatchResult:
    checklist = load_checklist(settings.checklists_path, tipo)
    work = work_dir or Path(tempfile.mkdtemp(prefix="batch_conf_"))
    rows: list[BatchRow] = []

    for idx, zip_path in enumerate(zip_paths, start=1):
        name = zip_path.name
        if on_progress:
            on_progress(f"[{idx}/{len(zip_paths)}] {name}")
        try:
            documents = load_from_zip(zip_path, work / f"extraidos_{idx}")
            if not documents:
                rows.append(
                    BatchRow(
                        ordem=idx,
                        zip_name=name,
                        beneficiario="—",
                        tipo=label_tipo(tipo),
                        atendidos=0,
                        parciais=0,
                        nao_atendidos=0,
                        situacao="ERRO — sem documentos",
                        cnpj="",
                        alertas="",
                        history_id="",
                        erro="ZIP sem arquivos legíveis",
                    )
                )
                continue

            rel = analisar_conformidade(settings, checklist, documents)
            cnpj_check = cross_check_cnpj(documents)
            # Anexa alertas CNPJ ao relatório
            if cnpj_check.alertas:
                rel.alertas = list(getattr(rel, "alertas", []) or []) + cnpj_check.alertas
                if cnpj_check.divergentes:
                    # marca itens de certidão/cnpj como parcial se CNPJ diverge
                    for item in rel.itens:
                        desc = item.descricao.lower()
                        if any(k in desc for k in ("cnpj", "fgts", "federal", "trabalhista", "ofício", "oficio")):
                            if item.status == StatusConformidade.ATENDIDO:
                                item.status = StatusConformidade.PARCIAL
                                item.motivo = (
                                    item.motivo
                                    + " CNPJ divergente entre documentos do pacote."
                                ).strip()
                                item.fonte = "regra+ml"

            counts = rel.contagem
            hid = ""
            if save_history:
                inv = [e.to_dict() for e in build_inventory(documents)]
                meta = save_analysis(
                    rel,
                    zip_name=name,
                    alertas=list(getattr(rel, "alertas", []) or []),
                    cnpj=cnpj_check.principal,
                    inventory=inv,
                )
                hid = meta.id

            rows.append(
                BatchRow(
                    ordem=idx,
                    zip_name=name,
                    beneficiario=rel.entidade_detectada,
                    tipo=label_tipo(tipo),
                    atendidos=counts["atendido"],
                    parciais=counts["parcial"],
                    nao_atendidos=counts["nao_atendido"],
                    situacao=_situacao(rel),
                    cnpj=cnpj_check.principal or "",
                    alertas=" | ".join(getattr(rel, "alertas", []) or [])[:500],
                    history_id=hid,
                )
            )
        except Exception as exc:
            rows.append(
                BatchRow(
                    ordem=idx,
                    zip_name=name,
                    beneficiario="—",
                    tipo=label_tipo(tipo),
                    atendidos=0,
                    parciais=0,
                    nao_atendidos=0,
                    situacao="ERRO",
                    cnpj="",
                    alertas="",
                    history_id="",
                    erro=str(exc)[:300],
                )
            )

    return BatchResult(
        rows=rows,
        generated_at=datetime.now().strftime("%d/%m/%Y %H:%M"),
    )
